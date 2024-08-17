from openpyxl import Workbook
filename = "material_list_canopy_mall"
def main():
    wb = Workbook()
    ws = wb.active
    ws['A1'] = "Block"
    ws['B1'] = "Total"
    ws['C1'] = "Missing"
    ws['D1'] = "Available"
    ws['E1'] = "Shulkers remaining"
    ws['F1'] = "Stacks remaining"
    ws['G1'] = "Items remaining"
    ws['H1'] = "Being gathered by"
    ws['I1'] = "Notes"
    file = open(filename + ".txt", "r")
    materialList = file.readlines()
    materialList = materialList[5:-3]
    file.close()
    for materialIndex in range(len(materialList)):
        materialList[materialIndex] = materialList[materialIndex].strip()
        material = materialList[materialIndex].split("|")
        ws.append([material[1].strip(), material[2].strip(), f"=MAX(0, ($B{materialIndex+2}-D{materialIndex+2}))", 0, f"=ROUNDDOWN(DIVIDE(C{materialIndex+2}, (27*64)))", f"=ROUNDDOWN(DIVIDE(C{materialIndex+2}-(E{materialIndex+2}*27*64), 64))", f"=MOD(C{materialIndex+2}, 64)"] )
    wb.save(filename + ".xlsx")

main()